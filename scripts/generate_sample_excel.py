#!/usr/bin/env python3
"""
Generate a sample Excel file that mirrors the Pod Resource Scanner Google Sheet layout.
Run from project root: python scripts/generate_sample_excel.py
Output: sample-pod-resource-scanner.xlsx (in current directory)
Use this to review structure and layout before pushing scanner changes.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

RUN_TS = "2026-02-04T120000Z"

# Light gray thin border for table cells
THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)
# Wrap text for long content (Reason, Action, Recommendations)
WRAP_ALIGN = Alignment(wrap_text=True, vertical="top")
# Zebra striping
ROW_FILL_LIGHT = PatternFill(start_color="F9F9F9", end_color="F9F9F9", fill_type="solid")


def style_header(cell, fill_hex="4472C4"):
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color=fill_hex, end_color=fill_hex, fill_type="solid")


def style_title(cell, fill_hex="B4C7E7"):
    cell.fill = PatternFill(start_color=fill_hex, end_color=fill_hex, fill_type="solid")


def apply_borders(ws, start_row, end_row, start_col, end_col):
    """Apply thin borders to a rectangular range."""
    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            ws.cell(row=r, column=c).border = THIN_BORDER


def write_dashboard(ws):
    ws.title = "Dashboard"
    # Row 1: title + Last updated
    ws["A1"] = "Pod Resource Scanner — Dashboard"
    ws["A1"].font = Font(bold=True, size=12, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="3D75B8", end_color="3D75B8", fill_type="solid")
    ws["E1"] = "Last updated: "
    ws["F1"] = f"Last scan: {RUN_TS}"
    ws.merge_cells("A1:D1")
    # Row 2: one-line explanation
    ws["A2"] = "Summary of the latest scan. Open a Run tab for full details and container-level recommendations."
    ws.merge_cells("A2:F2")
    ws["A2"].alignment = Alignment(wrap_text=True)
    # Rows 3–4: four numbers that tell the story (explainable)
    ws["A3"], ws["B3"] = "Pods", 54
    ws["C3"], ws["D3"] = "Containers", 59
    ws["E3"], ws["F3"] = "Nodes", 1
    ws["A4"], ws["B4"] = "Recommendations to review", 30
    for col in (2, 4, 6):
        ws.cell(row=3, column=col).alignment = Alignment(horizontal="right")
    ws.cell(row=4, column=2).alignment = Alignment(horizontal="right")
    # Rows 5–6: total CPU and Memory (requested vs allocatable, usage %)
    ws["A5"] = "CPU (cluster)"
    ws["B5"] = "requested: 3235 m"
    ws["C5"] = "allocatable: 4000 m"
    ws["D5"] = "usage: 80.9%"
    ws["A6"] = "Memory (cluster)"
    ws["B6"] = "requested: 4.73 Gi"
    ws["C6"] = "allocatable: 7.5 Gi"
    ws["D6"] = "usage: 63.1%"
    for r in (5, 6):
        for c in range(1, 5):
            ws.cell(row=r, column=c).fill = ROW_FILL_LIGHT
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 24


def write_run_sheet(wb):
    ws = wb.create_sheet(title=f"Run {RUN_TS}", index=1)
    # A1: Last scan
    ws["A1"] = f"Last scan: {RUN_TS}"
    ws["A1"].fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    # By Namespace (A2:C)
    ns_data = [
        ["By Namespace", "", ""],
        ["Namespace", "Pod Count", "Container Count"],
        ["hrbuddy", 8, 10],
        ["mcp", 4, 5],
        ["postgresql", 2, 2],
        ["kube-system", 6, 8],
        ["n8n", 1, 2],
        ["redis", 1, 1],
    ]
    for r, row in enumerate(ns_data, start=2):
        for c, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            if r >= 4 and (r - 4) % 2 == 1:
                cell.fill = ROW_FILL_LIGHT
    style_title(ws["A2"])
    style_header(ws["A3"]), style_header(ws["B3"]), style_header(ws["C3"])
    for r in range(4, 2 + len(ns_data)):
        ws.cell(row=r, column=2, value=ns_data[r - 2][1]).alignment = Alignment(horizontal="right")
        ws.cell(row=r, column=3, value=ns_data[r - 2][2]).alignment = Alignment(horizontal="right")
    apply_borders(ws, 2, 2 + len(ns_data) - 1, 1, 3)

    # Node utilization (E2:H)
    node_data = [
        ["Node utilization (%)", "", "", ""],
        ["Node", "CPU %", "Memory %", "Disk %"],
        ["k3s", 39.2, 15.1, 0],
    ]
    for r, row in enumerate(node_data, start=2):
        for c, val in enumerate(row, start=5):
            cell = ws.cell(row=r, column=c, value=val)
            if r >= 4 and (r - 4) % 2 == 1:
                cell.fill = ROW_FILL_LIGHT
    style_title(ws["E2"])
    for col in ["E", "F", "G", "H"]:
        style_header(ws[f"{col}3"])
    ws.cell(row=4, column=6, value=39.2).alignment = Alignment(horizontal="right")
    ws.cell(row=4, column=7, value=15.1).alignment = Alignment(horizontal="right")
    ws.cell(row=4, column=8, value=0).alignment = Alignment(horizontal="right")
    apply_borders(ws, 2, 2 + len(node_data) - 1, 5, 8)

    # Recommendations by type (J2:K)
    rec_type_data = [
        ["Recommendations by type", ""],
        ["Type", "Count"],
        ["change_limits", 28],
        ["scale_up", 1],
        ["scale_down", 1],
    ]
    for r, row in enumerate(rec_type_data, start=2):
        for c, val in enumerate(row, start=10):
            cell = ws.cell(row=r, column=c, value=val)
            if r >= 4 and (r - 4) % 2 == 1:
                cell.fill = ROW_FILL_LIGHT
    style_title(ws["J2"])
    style_header(ws["J3"]), style_header(ws["K3"])
    for r in range(4, 2 + len(rec_type_data)):
        ws.cell(row=r, column=11, value=rec_type_data[r - 2][1]).alignment = Alignment(horizontal="right")
    apply_borders(ws, 2, 2 + len(rec_type_data) - 1, 10, 11)

    # Recommendations (detailed) (L2:O) — wrap text, borders, zebra striping
    rec_detail_data = [
        ["Recommendations (detailed)", "", "", ""],
        ["Type", "Target", "Reason", "Action"],
        ["change_limits", "default/app-1/main", "Has requests but no limits set", "Set CPU/memory limits for predictability."],
        ["change_limits", "default/app-2/sidecar", "Memory limit >> request (limit 300 Mi, request 50 Mi)", "Consider lowering memory limit."],
        ["scale_up", "node:k3s", "High utilization: CPU 39%, memory 15%, disk 0%", "Consider adding nodes or moving workloads."],
    ]
    for r, row in enumerate(rec_detail_data, start=2):
        for c, val in enumerate(row, start=12):
            cell = ws.cell(row=r, column=c, value=val)
            if r >= 4 and c in (14, 15):  # Reason (N), Action (O)
                cell.alignment = WRAP_ALIGN
            if r >= 4 and (r - 4) % 2 == 1:
                cell.fill = ROW_FILL_LIGHT
    style_title(ws["L2"])
    for col in ["L", "M", "N", "O"]:
        style_header(ws[f"{col}3"])
    apply_borders(ws, 2, 2 + len(rec_detail_data) - 1, 12, 15)
    ws.column_dimensions["M"].width = 22
    ws.column_dimensions["N"].width = 48
    ws.column_dimensions["O"].width = 42

    # Container details (P1:W) with filter
    detail_header = [
        "Namespace", "Pod", "Container", "CPU Request", "CPU Limit",
        "Memory Request", "Memory Limit", "Recommendations",
    ]
    detail_data = [
        ["default", "app-1-abc", "main", "100m", "500m", "128Mi", "512Mi", "Set limits; CPU limit 4x+ request"],
        ["default", "app-1-abc", "sidecar", "50m", "200m", "64Mi", "256Mi", ""],
        ["hrbuddy", "hrbuddy-xyz", "hrbuddy", "200m", "1000m", "256Mi", "1Gi", "Memory limit >> request"],
        ["mcp", "mcp-server-1", "server", "400m", "2000m", "512Mi", "2Gi", ""],
        ["kube-system", "coredns-1", "coredns", "100m", "", "70Mi", "", "Has requests but no limits set"],
    ]
    ws["P1"] = "Container details (request / limit / suggestions)"
    style_title(ws["P1"])
    for c, h in enumerate(detail_header, start=16):
        cell = ws.cell(row=2, column=c, value=h)
        style_header(cell)
    for r, row in enumerate(detail_data, start=3):
        for c, val in enumerate(row, start=16):
            cell = ws.cell(row=r, column=c, value=val)
            if (r - 3) % 2 == 1:
                cell.fill = ROW_FILL_LIGHT
            if c == 23:  # Recommendations column (W)
                cell.alignment = WRAP_ALIGN
    apply_borders(ws, 1, 2 + len(detail_data), 16, 23)

    # Freeze top 2 rows and first 4 columns (freeze at E3)
    ws.freeze_panes = "E3"
    # Auto-filter on Container details (P2:W7)
    ws.auto_filter.ref = "P2:W7"
    # Column widths
    for col in range(1, 24):
        ws.column_dimensions[get_column_letter(col)].width = 14
    for col in range(16, 24):
        ws.column_dimensions[get_column_letter(col)].width = 18


def main():
    wb = Workbook()
    write_dashboard(wb.active)
    write_run_sheet(wb)
    import os
    out = "sample-pod-resource-scanner.xlsx"
    wb.save(out)
    path = os.path.abspath(out)
    print(f"Saved: {path}")
    print("Open in Excel to review layout before pushing.")


if __name__ == "__main__":
    main()
